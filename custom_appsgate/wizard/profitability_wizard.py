# -*- coding: utf-8 -*-

import base64
import io
from datetime import date

from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError

import logging
_logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, Alignment, PatternFill, Border, Side,
    )
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    _logger.warning(
        "openpyxl library not found. Install with: pip install openpyxl"
    )


class SaleProfitabilityWizard(models.TransientModel):
    _name = 'sale.profitability.wizard'
    _description = 'Sales Profitability Report Wizard'

    # ──────────────────────────────────────────────────────────────
    # FILTER FIELDS
    # ──────────────────────────────────────────────────────────────

    date_from = fields.Date(
        string='Date From',
        required=True,
        default=lambda self: date.today().replace(day=1),
    )
    date_to = fields.Date(
        string='Date To',
        required=True,
        default=lambda self: date.today(),
    )
    partner_ids = fields.Many2many(
        comodel_name='res.partner',
        relation='profitability_wiz_partner_rel',
        column1='wizard_id',
        column2='partner_id',
        string='Customers',
        domain=[('customer_rank', '>', 0)],
    )
    categ_ids = fields.Many2many(
        comodel_name='product.category',
        relation='profitability_wiz_categ_rel',
        column1='wizard_id',
        column2='categ_id',
        string='Product Categories',
    )
    product_ids = fields.Many2many(
        comodel_name='product.product',
        relation='profitability_wiz_product_rel',
        column1='wizard_id',
        column2='product_id',
        string='Products',
    )
    salesperson_ids = fields.Many2many(
        comodel_name='res.users',
        relation='profitability_wiz_user_rel',
        column1='wizard_id',
        column2='user_id',
        string='Salespersons',
    )
    company_id = fields.Many2one(
        comodel_name='res.company',
        string='Company',
        default=lambda self: self.env.company,
        required=True,
    )
    group_by = fields.Selection([
        ('order', 'By Order (Detail)'),
        ('product', 'By Product'),
        ('category', 'By Category'),
        ('customer', 'By Customer'),
    ], string='Group By', default='order', required=True)

    # ── Output Fields ──
    excel_file = fields.Binary(
        string='Excel File', readonly=True, attachment=False,
    )
    excel_filename = fields.Char(
        string='Filename', readonly=True,
    )

    # ── Result Lines ──
    line_ids = fields.One2many(
        'sale.profitability.wizard.line',
        'wizard_id',
        string='Report Lines',
    )

    # ── Summary Fields ──
    total_revenue = fields.Float(
        string='Total Revenue', readonly=True,
    )
    total_cost = fields.Float(
        string='Total Cost', readonly=True,
    )
    total_margin = fields.Float(
        string='Total Margin', readonly=True,
    )
    total_margin_percent = fields.Float(
        string='Overall Margin %', readonly=True,
    )
    record_count = fields.Integer(
        string='Records Found', readonly=True,
    )

    # ──────────────────────────────────────────────────────────────
    # CONSTRAINTS
    # ──────────────────────────────────────────────────────────────

    @api.constrains('date_from', 'date_to')
    def _check_dates(self):
        for wizard in self:
            if wizard.date_from and wizard.date_to:
                if wizard.date_from > wizard.date_to:
                    raise ValidationError(
                        _("'Date From' cannot be later than 'Date To'.")
                    )

    # ──────────────────────────────────────────────────────────────
    # DOMAIN BUILDER
    # ──────────────────────────────────────────────────────────────

    def _build_domain(self):
        self.ensure_one()
        domain = [
            ('order_date', '>=', self.date_from),
            ('order_date', '<=', self.date_to),
            ('company_id', '=', self.company_id.id),
        ]
        if self.partner_ids:
            domain.append(('partner_id', 'in', self.partner_ids.ids))
        if self.categ_ids:
            all_categ_ids = self.env['product.category'].search([
                ('id', 'child_of', self.categ_ids.ids)
            ]).ids
            domain.append(('categ_id', 'in', all_categ_ids))
        if self.product_ids:
            domain.append(('product_id', 'in', self.product_ids.ids))
        if self.salesperson_ids:
            domain.append(('user_id', 'in', self.salesperson_ids.ids))
        return domain

    # ──────────────────────────────────────────────────────────────
    # DATA RETRIEVAL
    # ──────────────────────────────────────────────────────────────

    def _get_report_data(self):
        self.ensure_one()
        domain = self._build_domain()
        records = self.env['sale.profitability.report'].search(
            domain, order='order_date desc, order_name'
        )
        data_lines = []
        for rec in records:
            data_lines.append({
                'order_name': rec.order_name or '',
                'order_date': rec.order_date,
                'partner_name': rec.partner_name or '',
                'product_name': rec.product_name or '',
                'category_name': rec.category_name or '',
                'quantity': rec.quantity,
                'revenue': rec.revenue,
                'cost': rec.cost,
                'margin': rec.margin,
                'margin_percent': rec.margin_percent,
            })
        return data_lines

    def _get_grouped_data(self):
        self.ensure_one()
        raw_data = self._get_report_data()

        if self.group_by == 'order':
            return raw_data

        group_field_map = {
            'product': 'product_name',
            'category': 'category_name',
            'customer': 'partner_name',
        }
        group_field = group_field_map[self.group_by]

        grouped = {}
        for line in raw_data:
            key = line.get(group_field) or 'Undefined'
            if key not in grouped:
                grouped[key] = {
                    'order_name': '',
                    'order_date': False,
                    'partner_name': key if self.group_by == 'customer' else '',
                    'product_name': key if self.group_by == 'product' else '',
                    'category_name': key if self.group_by == 'category' else '',
                    'quantity': 0.0,
                    'revenue': 0.0,
                    'cost': 0.0,
                    'margin': 0.0,
                }
            grp = grouped[key]
            grp['quantity'] += line['quantity']
            grp['revenue'] += line['revenue']
            grp['cost'] += line['cost']
            grp['margin'] += line['margin']

        result = []
        for key in sorted(grouped.keys()):
            vals = grouped[key]
            vals['margin_percent'] = (
                round((vals['margin'] / vals['revenue']) * 100, 2)
                if vals['revenue'] else 0.0
            )
            result.append(vals)

        return result

    # ──────────────────────────────────────────────────────────────
    # POPULATE WIZARD LINES
    # ──────────────────────────────────────────────────────────────

    def _populate_lines(self):
        self.ensure_one()
        self.line_ids.unlink()

        data = self._get_grouped_data()

        if not data:
            raise UserError(_(
                "No data found for the selected filters.\n\n"
                "Date range: %(date_from)s to %(date_to)s\n"
                "Company: %(company)s",
                date_from=self.date_from,
                date_to=self.date_to,
                company=self.company_id.name,
            ))

        lines_vals = []
        total_revenue = 0.0
        total_cost = 0.0
        total_margin = 0.0

        for line_data in data:
            total_revenue += line_data['revenue']
            total_cost += line_data['cost']
            total_margin += line_data['margin']

            lines_vals.append((0, 0, {
                'order_name': line_data.get('order_name', ''),
                'order_date': line_data.get('order_date'),
                'partner_name': line_data.get('partner_name', ''),
                'product_name': line_data.get('product_name', ''),
                'category_name': line_data.get('category_name', ''),
                'quantity': line_data.get('quantity', 0.0),
                'revenue': line_data.get('revenue', 0.0),
                'cost': line_data.get('cost', 0.0),
                'margin': line_data.get('margin', 0.0),
                'margin_percent': line_data.get('margin_percent', 0.0),
            }))

        overall_margin_pct = (
            round((total_margin / total_revenue) * 100, 2)
            if total_revenue else 0.0
        )

        self.write({
            'line_ids': lines_vals,
            'total_revenue': round(total_revenue, 2),
            'total_cost': round(total_cost, 2),
            'total_margin': round(total_margin, 2),
            'total_margin_percent': overall_margin_pct,
            'record_count': len(data),
        })

    # ──────────────────────────────────────────────────────────────
    # ACTIONS
    # ──────────────────────────────────────────────────────────────

    def action_preview(self):
        self.ensure_one()
        self._populate_lines()
        return {
            'type': 'ir.actions.act_window',
            'name': _('Sales Profitability Report'),
            'res_model': 'sale.profitability.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_print_pdf(self):
        self.ensure_one()
        self._populate_lines()
        return self.env.ref(
            'custom_appsgate.action_report_profitability_pdf'
        ).report_action(self)

    def action_export_excel(self):
        self.ensure_one()

        if not HAS_OPENPYXL:
            raise UserError(_(
                "The 'openpyxl' Python library is required.\n"
                "Install: pip install openpyxl"
            ))

        data = self._get_grouped_data()

        if not data:
            raise UserError(_(
                "No data found for the selected filters."
            ))

        wb = Workbook()
        ws = wb.active
        ws.title = 'Sales Profitability'

        # ── Styles ──
        title_font = Font(name='Calibri', bold=True, size=16, color='1F4E79')
        subtitle_font = Font(name='Calibri', italic=True,
                             size=10, color='666666')
        header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(
            start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True)
        data_font = Font(name='Calibri', size=10)
        number_alignment = Alignment(horizontal='right', vertical='center')
        text_alignment = Alignment(horizontal='left', vertical='center')
        center_alignment = Alignment(horizontal='center', vertical='center')
        positive_fill = PatternFill(
            start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        negative_fill = PatternFill(
            start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
        alt_row_fill = PatternFill(
            start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')
        totals_fill = PatternFill(
            start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
        totals_font = Font(name='Calibri', bold=True, size=11, color='1F4E79')

        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC'),
        )
        thick_bottom = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='medium', color='1F4E79'),
        )

        number_fmt = '#,##0.00'
        percent_fmt = '0.00'

        # ── Row 1: Title ──
        ws.merge_cells('A1:J1')
        ws['A1'].value = 'SALES PROFITABILITY REPORT'
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 30

        # ── Row 2: Period ──
        ws.merge_cells('A2:J2')
        ws['A2'].value = (
            f"Period: {self.date_from.strftime('%d %B %Y')} to "
            f"{self.date_to.strftime('%d %B %Y')}  |  "
            f"Company: {self.company_id.name}"
        )
        ws['A2'].font = subtitle_font
        ws['A2'].alignment = Alignment(horizontal='center')

        # ── Row 3: Filters ──
        filter_parts = []
        if self.partner_ids:
            filter_parts.append(
                f"Customers: {', '.join(self.partner_ids.mapped('name'))}")
        if self.categ_ids:
            filter_parts.append(
                f"Categories: {', '.join(self.categ_ids.mapped('complete_name'))}")
        if self.product_ids:
            filter_parts.append(
                f"Products: {', '.join(self.product_ids.mapped('display_name'))}")
        if self.salesperson_ids:
            filter_parts.append(
                f"Salespersons: {', '.join(self.salesperson_ids.mapped('name'))}")

        if filter_parts:
            ws.merge_cells('A3:J3')
            ws['A3'].value = 'Filters: ' + '  |  '.join(filter_parts)
            ws['A3'].font = subtitle_font
            ws['A3'].alignment = Alignment(horizontal='center')

        # ── Row 5: Headers ──
        headers = [
            ('#', 5),
            ('Order Ref', 18),
            ('Order Date', 14),
            ('Customer', 25),
            ('Product', 30),
            ('Category', 22),
            ('Qty', 12),
            ('Revenue', 16),
            ('Cost', 16),
            ('Margin', 16),
        ]

        header_row = 5
        ws.row_dimensions[header_row].height = 25

        for col_idx, (name, width) in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # ── Data Rows ──
        total_revenue = 0.0
        total_cost = 0.0
        total_margin = 0.0

        for row_offset, line in enumerate(data):
            row_idx = header_row + 1 + row_offset
            is_odd = row_offset % 2 == 1

            total_revenue += line.get('revenue', 0)
            total_cost += line.get('cost', 0)
            total_margin += line.get('margin', 0)

            values = [
                (row_offset + 1, center_alignment, None),
                (line.get('order_name', ''), text_alignment, None),
                (
                    line['order_date'].strftime('%d/%m/%Y')
                    if line.get('order_date') else '',
                    center_alignment, None,
                ),
                (line.get('partner_name', ''), text_alignment, None),
                (line.get('product_name', ''), text_alignment, None),
                (line.get('category_name', ''), text_alignment, None),
                (line.get('quantity', 0), number_alignment, number_fmt),
                (line.get('revenue', 0), number_alignment, number_fmt),
                (line.get('cost', 0), number_alignment, number_fmt),
                (line.get('margin', 0), number_alignment, number_fmt),
            ]

            for col_idx, (value, align, fmt) in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = align
                cell.border = thin_border
                if fmt:
                    cell.number_format = fmt
                if is_odd and col_idx <= 6:
                    cell.fill = alt_row_fill

            # Color margin column
            margin_val = line.get('margin', 0)
            margin_cell = ws.cell(row=row_idx, column=10)
            if margin_val >= 0:
                margin_cell.fill = positive_fill
                margin_cell.font = Font(
                    name='Calibri', size=10, color='2E7D32')
            else:
                margin_cell.fill = negative_fill
                margin_cell.font = Font(
                    name='Calibri', size=10, color='C62828')

        # ── Totals Row ──
        total_row = header_row + 1 + len(data)
        ws.row_dimensions[total_row].height = 25

        ws.merge_cells(
            start_row=total_row, start_column=1,
            end_row=total_row, end_column=6,
        )
        label_cell = ws.cell(row=total_row, column=1, value='TOTALS')
        label_cell.font = totals_font
        label_cell.fill = totals_fill
        label_cell.alignment = Alignment(horizontal='right')
        label_cell.border = thick_bottom

        for col in range(2, 7):
            c = ws.cell(row=total_row, column=col)
            c.fill = totals_fill
            c.border = thick_bottom

        total_values = [
            ('', None),
            (total_revenue, number_fmt),
            (total_cost, number_fmt),
            (total_margin, number_fmt),
        ]

        for col_offset, (value, fmt) in enumerate(total_values):
            col_idx = 7 + col_offset
            cell = ws.cell(row=total_row, column=col_idx, value=value)
            cell.font = totals_font
            cell.fill = totals_fill
            cell.alignment = number_alignment
            cell.border = thick_bottom
            if fmt:
                cell.number_format = fmt

        # ── Freeze & Filter ──
        ws.freeze_panes = f'A{header_row + 1}'
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f'A{header_row}:{last_col}{total_row - 1}'

        # ── Save ──
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = (
            f"sales_profitability_"
            f"{self.date_from.strftime('%Y%m%d')}_"
            f"{self.date_to.strftime('%Y%m%d')}.xlsx"
        )

        self.write({
            'excel_file': base64.b64encode(output.read()),
            'excel_filename': filename,
        })
        output.close()

        return {
            'type': 'ir.actions.act_window',
            'name': _('Download Excel Report'),
            'res_model': 'sale.profitability.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_clear(self):
        self.ensure_one()
        self.line_ids.unlink()
        self.write({
            'excel_file': False,
            'excel_filename': False,
            'total_revenue': 0.0,
            'total_cost': 0.0,
            'total_margin': 0.0,
            'total_margin_percent': 0.0,
            'record_count': 0,
        })
        return {
            'type': 'ir.actions.act_window',
            'name': _('Sales Profitability Report'),
            'res_model': 'sale.profitability.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }


class SaleProfitabilityWizardLine(models.TransientModel):
    _name = 'sale.profitability.wizard.line'
    _description = 'Sales Profitability Report Line'
    _order = 'id'

    wizard_id = fields.Many2one(
        'sale.profitability.wizard',
        string='Wizard',
        ondelete='cascade',
        required=True,
    )
    order_name = fields.Char(string='Order Ref')
    order_date = fields.Date(string='Order Date')
    partner_name = fields.Char(string='Customer')
    product_name = fields.Char(string='Product')
    category_name = fields.Char(string='Category')
    quantity = fields.Float(string='Quantity')
    revenue = fields.Float(string='Revenue')
    cost = fields.Float(string='Cost')
    margin = fields.Float(string='Margin')
    margin_percent = fields.Float(string='Margin %')
